import openpyxl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator

# Replace the file path with your actual file path
file_path = "/Users/abasaltbahrami/Downloads/MyographExperiment.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet_names = workbook.sheetnames

print("Sheet Names:")
for name in sheet_names:
    print(name)

# Verify if "190723" is present in the sheet names using partial match and case-insensitive search
sheet_name = "190723"
found_sheet = None
for name in sheet_names:
    if sheet_name.lower() in name.lower():
        found_sheet = workbook[name]
        print(f"Found sheet '{name}'")
        break

if found_sheet is None:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
else:
    sheet = found_sheet

# Extract data from the sheet and plot
x_data = [str(row[0] + 50)
          for row in sheet.iter_rows(min_row=2, values_only=True)]
y_data = [row[1] for row in sheet.iter_rows(min_row=2, values_only=True)]

# Create the plot
plt.plot(x_data, y_data, marker='o', linestyle='-')
plt.xlabel('X Data')
plt.ylabel('g/mg 3mM')
plt.title('Plot of g/mg 3mM vs X Data')
plt.grid(True)
plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
plt.show()
