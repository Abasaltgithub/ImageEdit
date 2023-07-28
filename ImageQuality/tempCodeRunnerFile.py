import openpyxl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator

# Replace the file path with your actual file path
file_path = "/Users/abasaltbahrami/Downloads/MyographExperiment.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet_names = workbook.sheetnames

print("Sheet Names:")
for name in sheet_names:
    print(name)

# Verify if "190723" is present in the sheet names using partial match and case-insensitive search
sheet_name = "190723"
found_sheet = None
for name in sheet_names:
    if sheet_name.lower() in name.lower():
        found_sheet = workbook[name]
        print(f"Found sheet '{name}'")
        break

if found_sheet is None:
    print(f"Sheet '{sheet_name}' not found in the workbook.")
else:
    sheet = found_sheet

    # Look for the columns "Strech" and "g/mg 3mM" and collect their values
    strech_column_index = None
    g_mg_column_index = None
    for column_index, column_cells in enumerate(sheet.iter_cols()):
        if any(cell.value and "Strech" in str(cell.value) for cell in column_cells):
            strech_column_index = column_index + 1
        if any(cell.value and "g/mg 3mM" in str(cell.value) for cell in column_cells):
            g_mg_column_index = column_index + 1

        if strech_column_index is not None and g_mg_column_index is not None:
            break

    if strech_column_index is None or g_mg_column_index is None:
        print("Columns 'Strech' and 'g/mg 3mM' not found in the sheet.")
    else:
        # Collect the values for the "Strech" and "g/mg 3mM" columns
        strech_values = [row[strech_column_index - 1] for row in sheet.iter_rows(
            values_only=True) if row[strech_column_index - 1] is not None]
        g_mg_values = [row[g_mg_column_index - 1] for row in sheet.iter_rows(
            values_only=True) if row[g_mg_column_index - 1] is not None]

        # Plot "g/mg 3mM" versus "Strech"
        plt.figure(figsize=(10, 6))  # Adjust the figure size as needed
        plt.plot(strech_values, g_mg_values, marker='o', linestyle='-')
        plt.xlabel("Strech")
        plt.ylabel("g/mg 3mM")
        plt.title("g/mg 3mM versus Strech")
        plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
        plt.grid(True)

        # Set the number of ticks on the x-axis and y-axis
        ax = plt.gca()
        ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))

        plt.tight_layout()  # Automatically adjust padding
        plt.show()
